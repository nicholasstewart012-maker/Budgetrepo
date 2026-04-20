from __future__ import annotations

import mimetypes
import os
import re
import csv
import time
from datetime import datetime
from functools import lru_cache
from pathlib import Path

from docx import Document as DocxDocument
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from docx.table import Table
from docx.text.paragraph import Paragraph
from fastapi import HTTPException
from fastapi.responses import FileResponse
from urllib.parse import quote

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional dependency
    load_workbook = None


INTAKE_FOLDER = Path(os.getenv("INTAKE_FOLDER", r"C:\Spark\intake")).resolve()
VIEWER_DEBUG = os.getenv("SPARK_VIEWER_DEBUG", "false").lower() == "true"


def _viewer_debug(message: str, **data) -> None:
    if not VIEWER_DEBUG:
        return
    detail = " ".join(f"{k}={v!r}" for k, v in data.items())
    print(f"[Spark Viewer] {message} {detail}".strip())


# Performance logging: always-on (cheap) so we can spot regressions in prod logs
# without needing to flip SPARK_VIEWER_DEBUG.
SLOW_ENDPOINT_MS = int(os.getenv("SPARK_VIEWER_SLOW_MS", "150"))


def _perf_log(label: str, start: float, **extra) -> None:
    elapsed_ms = int((time.perf_counter() - start) * 1000)
    if elapsed_ms >= SLOW_ENDPOINT_MS:
        extra_str = " ".join(f"{k}={v!r}" for k, v in extra.items())
        print(f"[Spark Viewer PERF] {label} took {elapsed_ms}ms {extra_str}".strip())


@lru_cache(maxsize=256)
def _pdf_page_count_fast(path_str: str, mtime_ns: int, file_size: int) -> int:
    """
    Cheap PDF page count — opens the PDF just far enough to read the page
    tree root, without extracting any text.  Used by /document/meta where
    we only need the page count, not the text content of every page.

    Cached keyed on (path, mtime_ns, size) so repeat hits on the same
    unchanged file are essentially free.
    """
    from pypdf import PdfReader
    try:
        reader = PdfReader(path_str)
        return len(reader.pages)
    except Exception:
        return 0


def resolve_document_path(path: str) -> Path:
    if not path or not str(path).strip():
        raise HTTPException(status_code=400, detail="path is required")

    requested = Path(str(path).strip()).expanduser()
    file_path = requested if requested.is_absolute() else (INTAKE_FOLDER / requested)
    file_path = file_path.resolve()
    try:
        file_path.relative_to(INTAKE_FOLDER)
    except ValueError as exc:
        raise HTTPException(status_code=403, detail="Access denied: path is outside the intake folder") from exc

    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(status_code=404, detail=f"File not found: {path}")

    return file_path


def _normalize_text(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip()).lower()


def _split_sentences(text: str) -> list[str]:
    if not text:
        return []
    cleaned = re.sub(r"\s+", " ", text).strip()
    return [part.strip() for part in re.split(r"(?<=[.!?])\s+", cleaned) if part.strip()]


def _clean_preview_text(text: str) -> str:
    text = (text or "").replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]{2,}", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _xlsx_col_letter(col_idx: int) -> str:
    letters = ""
    value = max(int(col_idx), 1)
    while value > 0:
        value, remainder = divmod(value - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _xlsx_display_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat(sep=" ", timespec="seconds")
        except TypeError:
            return value.isoformat()
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.15g}"
    return str(value).strip()


def _xlsx_range_ref(min_row: int, max_row: int, min_col: int, max_col: int) -> str:
    return f"{_xlsx_col_letter(min_col)}{min_row}:{_xlsx_col_letter(max_col)}{max_row}"


def _xlsx_sheet_payload(file_path: Path) -> dict:
    if load_workbook is None:
        return {"sheets": [], "extraction_status": "openpyxl_unavailable"}

    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as exc:
        return {"sheets": [], "extraction_status": f"failed:{exc}"}

    sheets: list[dict] = []
    try:
        for sheet in workbook.worksheets:
            if getattr(sheet, "sheet_state", "visible") != "visible":
                continue

            try:
                dims = sheet.calculate_dimension()
            except Exception:
                dims = ""
            if not dims or ":" not in dims:
                sheets.append({
                    "sheet_name": sheet.title,
                    "used_range": None,
                    "max_row": int(getattr(sheet, "max_row", 0) or 0),
                    "max_column": int(getattr(sheet, "max_column", 0) or 0),
                    "headers": [],
                    "rows": [],
                    "preview_rows": [],
                })
                continue

            from openpyxl.utils.cell import range_boundaries
            min_col, min_row, max_col, max_row = range_boundaries(dims)
            rows: list[list[str]] = []
            for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                values = [_xlsx_display_value(cell.value if cell is not None else None) for cell in row]
                if any(values):
                    rows.append(values)

            if not rows:
                sheets.append({
                    "sheet_name": sheet.title,
                    "used_range": _xlsx_range_ref(min_row, max_row, min_col, max_col),
                    "max_row": int(getattr(sheet, "max_row", 0) or 0),
                    "max_column": int(getattr(sheet, "max_column", 0) or 0),
                    "headers": [],
                    "rows": [],
                    "preview_rows": [],
                })
                continue

            header_row = 0
            header_candidates = rows[:5]
            for idx, row in enumerate(header_candidates):
                text_cells = sum(1 for value in row if value and not re.fullmatch(r"[-+]?\d+(?:\.\d+)?", value))
                if text_cells >= 2:
                    header_row = idx
                    break

            headers = [value or f"Column {_xlsx_col_letter(i + min_col)}" for i, value in enumerate(rows[header_row])]
            preview_rows = []
            for row_number, row_values in enumerate(rows[header_row + 1:header_row + 9], start=min_row + header_row + 1):
                preview_rows.append({
                    "row_number": row_number,
                    "cells": [
                        {
                            "column": _xlsx_col_letter(min_col + idx),
                            "header": headers[idx] if idx < len(headers) else f"Column {_xlsx_col_letter(min_col + idx)}",
                            "value": row_values[idx] if idx < len(row_values) else "",
                        }
                        for idx in range(max(len(headers), len(row_values)))
                    ],
                })

            sheets.append({
                "sheet_name": sheet.title,
                "used_range": _xlsx_range_ref(min_row, max_row, min_col, max_col),
                "max_row": int(getattr(sheet, "max_row", 0) or max_row),
                "max_column": int(getattr(sheet, "max_column", 0) or max_col),
                "headers": headers,
                "rows": rows,
                "preview_rows": preview_rows,
            })
    finally:
        workbook.close()

    return {"sheets": sheets, "extraction_status": "ok"}


def _detect_csv_dialect(sample: str) -> csv.Dialect:
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except Exception:
        class _FallbackDialect(csv.Dialect):
            delimiter = ","
            quotechar = '"'
            doublequote = True
            skipinitialspace = True
            lineterminator = "\n"
            quoting = csv.QUOTE_MINIMAL
        return _FallbackDialect()


def _csv_sheet_payload(file_path: Path) -> dict:
    try:
        raw = file_path.read_text(encoding="utf-8-sig", errors="ignore")
    except Exception as exc:
        return {"sheets": [], "extraction_status": f"failed:{exc}"}

    if not raw.strip():
        return {"sheets": [], "extraction_status": "empty"}

    lines = raw.splitlines()
    sample = "\n".join(lines[:20]) if lines else raw[:4096]
    dialect = _detect_csv_dialect(sample)
    try:
        reader = csv.reader(lines, dialect=dialect)
        rows = [[str(cell).strip() for cell in row] for row in reader]
    except Exception as exc:
        return {"sheets": [], "extraction_status": f"failed:{exc}"}

    rows = [row for row in rows if any((cell or "").strip() for cell in row)]
    if not rows:
        return {"sheets": [], "extraction_status": "empty"}

    max_col = max((len(row) for row in rows), default=0)
    if max_col <= 0:
        return {"sheets": [], "extraction_status": "empty"}
    padded_rows = [row + [""] * (max_col - len(row)) for row in rows]

    header_row = 0
    for idx, row in enumerate(padded_rows[:5]):
        text_cells = sum(1 for value in row if value and not re.fullmatch(r"[-+]?\d+(?:\.\d+)?", value))
        if text_cells >= 2:
            header_row = idx
            break

    headers = [value or f"Column {_xlsx_col_letter(i + 1)}" for i, value in enumerate(padded_rows[header_row])]
    preview_rows = []
    for row_number, row_values in enumerate(padded_rows[header_row + 1:header_row + 9], start=header_row + 2):
        preview_rows.append({
            "row_number": row_number,
            "cells": [
                {
                    "column": _xlsx_col_letter(idx + 1),
                    "header": headers[idx] if idx < len(headers) else f"Column {_xlsx_col_letter(idx + 1)}",
                    "value": row_values[idx] if idx < len(row_values) else "",
                }
                for idx in range(max(len(headers), len(row_values)))
            ],
        })

    sheets = [{
        "sheet_name": file_path.stem or "Sheet1",
        "used_range": _xlsx_range_ref(1, len(padded_rows), 1, max_col),
        "max_row": len(padded_rows),
        "max_column": max_col,
        "headers": headers,
        "rows": padded_rows,
        "preview_rows": preview_rows,
    }]
    return {"sheets": sheets, "extraction_status": "ok"}


def _normalize_content_type(ext: str, fallback: str = "application/octet-stream") -> str:
    return {
        "pdf": "application/pdf",
        "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "xlsm": "application/vnd.ms-excel.sheet.macroEnabled.12",
        "csv": "text/csv",
        "txt": "text/plain",
        "md": "text/markdown",
        "rtf": "application/rtf",
    }.get(ext, fallback)


def _split_paragraph_groups(text: str) -> list[list[str]]:
    groups: list[list[str]] = []
    current: list[str] = []
    for raw_line in (text or "").replace("\r\n", "\n").replace("\r", "\n").split("\n"):
        line = raw_line.rstrip()
        if line.strip():
            current.append(line)
        elif current:
            groups.append(current)
            current = []
    if current:
        groups.append(current)
    return groups


def _looks_like_heading(line: str) -> bool:
    stripped = line.strip()
    if not stripped or len(stripped) > 90:
        return False
    if stripped.endswith(":"):
        return True
    words = stripped.split()
    if len(words) <= 10 and stripped.upper() == stripped:
        return True
    alpha = [w for w in words if re.search(r"[A-Za-z]", w)]
    if 0 < len(alpha) <= 8 and all(w[:1].isupper() for w in alpha):
        return True
    return False


def _is_structured_line(line: str) -> bool:
    stripped = line.strip()
    return bool(
        re.match(r"^(?:[-*•]|(\d+[\.)]))\s+", stripped)
        or "|" in stripped
        or re.match(r"^[A-Z][A-Z0-9 _/-]{3,}$", stripped)
    )


def _is_markdown_heading(line: str) -> tuple[bool, int | None]:
    match = re.match(r"^(#{1,6})\s+(.*)$", line.strip())
    if not match:
        return False, None
    return True, min(len(match.group(1)), 3)


def _is_markdown_list_item(line: str) -> bool:
    return bool(re.match(r"^(?:[-*+]|(?:\d+[\.)]))\s+", line.strip()))


def _is_markdown_table_group(lines: list[str]) -> bool:
    if len(lines) < 2 or not any("|" in line for line in lines):
        return False
    separator_found = any(re.match(r"^\s*\|?\s*[:\-\| ]+\|?\s*$", line) for line in lines)
    return separator_found or all("|" in line for line in lines)


def _split_markdown_table_row(line: str) -> list[str]:
    cells = [cell.strip() for cell in re.split(r"(?<!\\)\|", line.strip())]
    if cells and cells[0] == "":
        cells = cells[1:]
    if cells and cells[-1] == "":
        cells = cells[:-1]
    return [cell.replace("\\|", "|").strip() for cell in cells if cell is not None]


def _collect_candidates(evidence_anchor: str, evidence_context: str, evidence_text: str, snippet: str, chunk_text: str, answer: str, question: str) -> list[dict]:
    candidates: list[dict] = []
    seen: set[tuple[str, str]] = set()

    def add(match_type: str, phrase: str, score: float) -> None:
        normalized = _normalize_text(phrase)
        if not normalized:
            return
        key = (match_type, normalized)
        if key in seen:
            return
        seen.add(key)
        candidates.append({
            "matchType": match_type,
            "phrase": phrase.strip(),
            "score": score,
        })

    def add_windows(text: str, match_type: str, base_score: float) -> None:
        if not text:
            return
        add(match_type, text, base_score)
        for sentence in _split_sentences(text):
            add(match_type, sentence, base_score - 0.05)
            words = re.findall(r"\b[\w'-]+\b", _normalize_text(sentence))
            for size in (12, 10, 8, 6, 5, 4, 3):
                if len(words) < size:
                    continue
                for idx in range(len(words) - size + 1):
                    phrase = " ".join(words[idx:idx + size]).strip()
                    add(match_type, phrase, base_score - 0.1)

    add_windows(evidence_anchor, "evidenceAnchor", 1.14)
    add_windows(evidence_context, "evidenceContext", 1.1)
    add_windows(evidence_text, "evidenceText", 1.05)
    add_windows(snippet, "snippet", 1.0)
    add_windows(chunk_text, "chunkText", 0.9)
    add_windows(answer, "answerPhrase", 0.8)

    if question and question.strip():
        add("manual", question, 0.6)
        words = re.findall(r"\b[\w'-]+\b", _normalize_text(question))
        for size in (6, 5, 4, 3):
            if len(words) < size:
                continue
            for idx in range(len(words) - size + 1):
                phrase = " ".join(words[idx:idx + size]).strip()
                add("manual", phrase, 0.58)

    return candidates


NON_PDF_BLOCK_LIMIT = 3
NON_PDF_PRIMARY_LIMIT = 1
NON_PDF_INLINE_PER_BLOCK_LIMIT = 2
NON_PDF_TOTAL_INLINE_LIMIT = 5
NON_PDF_MIN_PHRASE_LEN = 20

COMMON_PHRASE_TOKENS = {
    "the", "and", "for", "are", "that", "with", "this", "from", "have", "has", "you", "your", "was", "were",
    "will", "shall", "must", "may", "can", "not", "all", "any", "code", "policy", "employee", "employees",
    "report", "reports", "section", "document", "text", "item", "items", "list", "table", "data",
    "source", "answer", "paragraph", "file", "block", "open",
}

SOURCE_MATCH_PRIORITY = {
    "evidenceAnchor": 7,
    "evidenceContext": 6,
    "evidenceText": 5,
    "snippet": 4,
    "chunkText": 3,
    "answerPhrase": 2,
    "question": 1,
}

PDF_MATCH_LIMIT = 5
PDF_MIN_MEANINGFUL_TEXT_LEN = 30


def _tokenize_normalized(text: str) -> list[str]:
    return re.findall(r"\b[\w'-]+\b", text or "")


def _meaningful_token_set(text: str) -> set[str]:
    return {
        token
        for token in _tokenize_normalized(_normalize_text(text))
        if len(token) > 2 and token not in COMMON_PHRASE_TOKENS
    }


def _strip_citations(text: str) -> str:
    return re.sub(r"\s*\[[^\]]+\]", "", text or "").strip()


def _looks_like_pdf_heading(text: str) -> bool:
    cleaned = re.sub(r"\s+", " ", (text or "").strip())
    if not cleaned:
        return True
    tokens = _tokenize_normalized(_normalize_text(cleaned))
    if not tokens:
        return True
    if len(tokens) > 8 or len(cleaned) > 80:
        return False
    section_terms = {
        "purpose",
        "scope",
        "policy",
        "overview",
        "definitions",
        "introduction",
        "revision",
        "history",
        "procedure",
        "procedures",
        "contents",
        "statement",
        "guide",
        "applicability",
        "responsibilities",
        "approval",
        "committee",
        "sponsor",
        "owner",
        "effective",
        "background",
        "references",
        "appendix",
        "objectives",
        "summary",
        "roles",
    }
    starts_numbered = bool(re.match(r"^\d+(?:\.\d+)*\s+", cleaned))
    has_label_punct = bool(re.search(r"[,;:]", cleaned))
    alpha = re.sub(r"[^A-Za-z]", "", cleaned)
    upper_ratio = (len(re.sub(r"[^A-Z]", "", cleaned)) / len(alpha)) if alpha else 0
    words = re.findall(r"\b[A-Za-z][A-Za-z'-]*\b", cleaned)
    titlecase_ratio = (
        sum(1 for word in words if word[:1].isupper() and word[1:].islower()) / len(words)
    ) if words else 0
    has_section_term = any(token in section_terms for token in tokens)
    title_like = not has_label_punct and titlecase_ratio >= 0.65 and len(tokens) <= 8
    return (
        (starts_numbered and (has_section_term or len(tokens) <= 4))
        or (upper_ratio >= 0.72 and len(tokens) <= 5)
        or (has_section_term and len(tokens) <= 4 and not has_label_punct)
        or title_like
    )


def _looks_like_body_text(text: str) -> bool:
    cleaned = re.sub(r"\s+", " ", (text or "").strip())
    if len(cleaned) < PDF_MIN_MEANINGFUL_TEXT_LEN:
        return False
    if _looks_like_pdf_heading(cleaned):
        return False
    tokens = _meaningful_token_set(cleaned)
    return len(tokens) >= 4 and (bool(re.search(r"[,.;:]", cleaned)) or len(cleaned) >= 70)


def _looks_like_pdf_metadata_block(text: str) -> bool:
    cleaned = re.sub(r"\s+", " ", (text or "").strip())
    # Lowered minimum length — short metadata labels still poison scoring
    if len(cleaned) < 35:
        return False
    lower = cleaned.lower()
    label_terms = (
        "sponsoring department",
        "executive sponsor",
        "board committee",
        "committee approval date",
        "policy statement",
        "change management policy",
        "effective date",
        "approved by",
        "document owner",
        "policy owner",
        "version number",
        "revision date",
        "review date",
        "next review",
        "last reviewed",
        "last modified",
        "classification",
        "document number",
        "applicable to",
    )
    label_hits = sum(1 for term in label_terms if term in lower)
    colon_count = cleaned.count(":")
    if label_hits >= 2:
        return True
    if colon_count >= 3 and not re.search(r"[.!?]", cleaned):
        return True
    return False


def _looks_like_revision_table(text: str) -> bool:
    """Returns True when the text block looks like a revision-history table.

    Revision tables have a high density of:
    - Date-like patterns (MM/DD/YYYY, month-year)
    - Version number patterns (v1.0, 2.3)
    - Short fields separated by pipes or multiple colons
    """
    cleaned = re.sub(r"\s+", " ", (text or "").strip())
    if len(cleaned) < 30:
        return False
    date_count = len(re.findall(
        r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b|\b(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\w*\s+\d{4}\b",
        cleaned, re.IGNORECASE,
    ))
    version_count = len(re.findall(r"\bv?\d+\.\d+", cleaned, re.IGNORECASE))
    colon_count = cleaned.count(":")
    pipe_count = cleaned.count("|")
    if (date_count >= 2 or version_count >= 2) and (colon_count >= 2 or pipe_count >= 2):
        return True
    if date_count >= 3:
        return True
    if pipe_count >= 3:
        return True
    return False


def _looks_like_pdf_sentence(text: str) -> bool:
    """Returns True when the text looks like a real body sentence."""
    cleaned = re.sub(r"\s+", " ", (text or "").strip())
    if not cleaned or len(cleaned) < 40:
        return False
    if _looks_like_pdf_heading(cleaned):
        return False
    tokens = _tokenize_normalized(_normalize_text(cleaned))
    if len(tokens) < 8:
        return False
    has_punct = bool(re.search(r"[,.;:!?]", cleaned))
    has_mixed_case = bool(re.search(r"[a-z]", cleaned) and re.search(r"[A-Z]", cleaned))
    return has_punct or has_mixed_case


def _strip_leading_pdf_heading(text: str) -> str:
    cleaned = re.sub(r"\s+", " ", (text or "").strip())
    if not cleaned:
        return ""
    section = r"(?:PURPOSE|SCOPE|POLICY|OVERVIEW|DEFINITIONS|INTRODUCTION|REVISION HISTORY|PROCEDURE|CONTENTS|ROLES)"
    stripped = re.sub(rf"^\d+(?:\.\d+)*\s+{section}\s+(?=[A-Z][a-z])", "", cleaned, count=1)
    return stripped.strip() or cleaned


def _pdf_phrase_anchor(block_text: str, query: dict) -> str:
    query_phrase = _strip_leading_pdf_heading(_strip_citations(query.get("phrase", "")))
    block_phrase = _strip_leading_pdf_heading(block_text)
    if query_phrase and _looks_like_body_text(query_phrase):
        query_norm = _normalize_text(query_phrase)
        block_norm = _normalize_text(block_text)
        if query_norm and query_norm in block_norm:
            return query_phrase
    return block_phrase


def _pdf_text_blocks(page_text: str) -> list[str]:
    cleaned = _clean_preview_text(page_text or "")
    if not cleaned:
        return []

    blocks: list[str] = []
    for paragraph in re.split(r"\n\s*\n+", cleaned):
        paragraph = _clean_preview_text(paragraph)
        if not paragraph:
            continue
        blocks.append(paragraph)
        if len(paragraph) > 220:
            for sentence in _split_sentences(paragraph):
                sentence = _clean_preview_text(sentence)
                if sentence and len(sentence) >= PDF_MIN_MEANINGFUL_TEXT_LEN:
                    blocks.append(sentence)

    deduped: list[str] = []
    seen: set[str] = set()
    for block in blocks:
        key = _normalize_text(block)
        if not key or key in seen:
            continue
        seen.add(key)
        deduped.append(block)
    return deduped


def _collect_pdf_queries(evidence_anchor: str, evidence_context: str, evidence_text: str, snippet: str, chunk_text: str, answer: str, question: str) -> list[dict]:
    queries: list[dict] = []
    seen: set[str] = set()

    def add(match_type: str, text: str, score: float) -> None:
        phrase = _strip_citations(text)
        normalized = _normalize_text(phrase)
        if len(normalized) < 8:
            return
        key = normalized
        if key in seen:
            return
        seen.add(key)
        queries.append({
            "matchType": match_type,
            "phrase": phrase,
            "normalizedPhrase": normalized,
            "tokens": _meaningful_token_set(phrase),
            "score": score,
            "priority": SOURCE_MATCH_PRIORITY.get(match_type, 0),
        })

    def add_source(match_type: str, text: str, base_score: float) -> None:
        source = _strip_citations(text)
        if not source:
            return
        add(match_type, source, base_score)
        for sentence in _split_sentences(source):
            add(match_type, sentence, base_score - 0.03)

    add_source("evidenceAnchor", evidence_anchor, 1.18)
    add_source("evidenceContext", evidence_context, 1.14)
    add_source("evidenceText", evidence_text, 1.1)
    add_source("snippet", snippet, 1.04)
    add_source("chunkText", chunk_text, 0.96)
    add_source("answerPhrase", answer, 0.92)
    add_source("question", question, 0.62)

    return sorted(
        [query for query in queries if query["tokens"] or query["matchType"] != "question"],
        key=lambda query: (-float(query["score"]), -int(query["priority"]), -len(query["normalizedPhrase"])),
    )


def _score_pdf_block(query: dict, block_text: str, page_number: int, target_page: int | None, page_count: int) -> float:
    block_norm = _normalize_text(block_text)
    query_norm = query.get("normalizedPhrase", "")
    if not block_norm or not query_norm:
        return 0.0

    block_heading = _looks_like_pdf_heading(block_text)
    block_metadata = _looks_like_pdf_metadata_block(block_text)
    query_heading = _looks_like_pdf_heading(query.get("phrase", ""))
    block_body = _looks_like_body_text(block_text)
    query_body = _looks_like_body_text(query.get("phrase", ""))
    query_metadata = _looks_like_pdf_metadata_block(query.get("phrase", ""))
    query_tokens = set(query.get("tokens") or [])
    block_tokens = _meaningful_token_set(block_text)
    overlap = len(query_tokens & block_tokens)
    query_coverage = overlap / max(len(query_tokens), 1)
    block_density = overlap / max(len(block_tokens), 1)
    exact_bonus = 0.0

    if len(query_norm) >= 16 and query_norm in block_norm:
        exact_bonus = 0.55 if block_heading else 1.15
    elif len(block_norm) >= PDF_MIN_MEANINGFUL_TEXT_LEN and block_norm in query_norm:
        exact_bonus = 0.25 if block_heading else 0.72

    if exact_bonus <= 0 and query_coverage < 0.32:
        return 0.0
    if (
        exact_bonus <= 0
        and query.get("matchType") == "snippet"
        and not re.search(r"[,.;:]", query.get("phrase", ""))
        and query_coverage < 0.75
    ):
        return 0.0

    block_revision_table = _looks_like_revision_table(block_text)

    body_bonus = 0.32 if block_body else 0.0
    query_body_bonus = 0.14 if query_body else 0.0
    label_bonus = 0.16 if block_body and ":" in block_text else 0.0
    # Heading penalty raised from 1.05 → 1.35
    heading_penalty = 1.35 if block_heading else 0.0
    query_heading_penalty = 0.28 if query_heading and not query_body else 0.0
    snippet_heading_penalty = 0.36 if query.get("matchType") == "snippet" and block_heading else 0.0
    source_heading_penalty = 0.0
    if block_heading and not query_body:
        source_heading_penalty += 0.55
        if query.get("matchType") == "evidenceText":
            source_heading_penalty += 0.55
    # Metadata penalty raised from 1.35 → 1.80
    if block_metadata and query.get("matchType") in {"evidenceText", "snippet"}:
        source_heading_penalty += 1.80
    if block_heading and query_metadata and query.get("matchType") in {"evidenceText", "snippet"}:
        source_heading_penalty += 1.30
    # Revision table penalty
    revision_table_penalty = 1.20 if block_revision_table else 0.0
    short_penalty = 0.35 if len(block_norm) < PDF_MIN_MEANINGFUL_TEXT_LEN else 0.0
    length_bonus = min(0.45, len(block_norm) / 360.0)
    source_bonus = float(query.get("score", 0.0))
    priority_bonus = int(query.get("priority", 0)) * 0.035

    proximity_bonus = 0.0
    if target_page is not None:
        distance = abs(page_number - target_page)
        proximity_bonus = max(0.0, 0.18 - (distance / max(page_count, 1)) * 0.18)

    return (
        source_bonus
        + priority_bonus
        + exact_bonus
        + query_coverage * 1.1
        + block_density * 0.62
        + body_bonus
        + query_body_bonus
        + label_bonus
        + length_bonus
        + proximity_bonus
        - heading_penalty
        - query_heading_penalty
        - snippet_heading_penalty
        - source_heading_penalty
        - revision_table_penalty
        - short_penalty
    )


def _select_pdf_highlights(preview_blocks: list[dict], queries: list[dict], page_number: int | None = None) -> list[dict]:
    ranked: list[dict] = []
    page_count = len(preview_blocks)

    for idx, page in enumerate(preview_blocks):
        current_page = int(page.get("pageNumber", idx + 1) or idx + 1)
        page_text = _normalize_block_text(page)
        for block_text in _pdf_text_blocks(page_text):
            block_norm = _normalize_text(block_text)
            if not block_norm:
                continue
            for query in queries:
                score = _score_pdf_block(query, block_text, current_page, page_number, page_count)
                if score <= 0:
                    continue
                heading_like = _looks_like_pdf_heading(block_text)
                body_like = _looks_like_body_text(block_text)
                metadata_like = _looks_like_pdf_metadata_block(block_text)
                sentence_like = _looks_like_pdf_sentence(block_text)
                ranked.append({
                    "pageNumber": current_page,
                    "phrase": _pdf_phrase_anchor(block_text, query),
                    "matchType": query.get("matchType", "snippet"),
                    "score": round(score, 4),
                    "isPrimary": False,
                    # Rich metadata forwarded to frontend for re-ranking
                    "headingLike": heading_like,
                    "bodyLike": body_like,
                    "metadataLike": metadata_like,
                    "sentenceLike": sentence_like,
                    "phraseLength": len(_normalize_text(block_text)),
                })

    ranked.sort(key=lambda item: (-float(item["score"]), abs(int(item["pageNumber"]) - page_number) if page_number is not None else 0, -len(_normalize_text(item["phrase"]))))

    selected: list[dict] = []
    seen: set[tuple[int, str]] = set()
    for match in ranked:
        key = (int(match["pageNumber"]), _normalize_text(match["phrase"]))
        if key in seen:
            continue
        seen.add(key)
        match["isPrimary"] = len(selected) == 0
        selected.append(match)
        if len(selected) >= PDF_MATCH_LIMIT:
            break
    return selected


def _phrase_is_common(normalized_phrase: str) -> bool:
    tokens = _tokenize_normalized(normalized_phrase)
    if not tokens:
        return True
    if len(normalized_phrase) < 12:
        return True
    if len(tokens) <= 2 and all(token in COMMON_PHRASE_TOKENS for token in tokens):
        return True
    common_ratio = sum(1 for token in tokens if token in COMMON_PHRASE_TOKENS) / max(len(tokens), 1)
    return common_ratio >= 0.85


def _phrase_boundary_pattern(normalized_phrase: str) -> str:
    return rf"(?<![a-z0-9_]){re.escape(normalized_phrase)}(?![a-z0-9_])"


def _phrase_occurs_in_normalized_text(normalized_haystack: str, normalized_phrase: str) -> bool:
    """Return True only when a candidate phrase appears as a safe text span.

    Raw substring matching lets short candidates such as "IR" match inside
    unrelated words like "primary" and "first". Short / low-token phrases must
    therefore use token-boundary matching, while longer evidence phrases can
    still use substring matching for extraction resilience.
    """
    haystack = normalized_haystack or ""
    phrase = normalized_phrase or ""
    if not haystack or not phrase:
        return False

    tokens = _tokenize_normalized(phrase)
    meaningful = _meaningful_token_set(phrase)
    if not tokens:
        return False

    if len(tokens) == 1:
        token = tokens[0]
        if len(token) <= 2 or token in COMMON_PHRASE_TOKENS:
            return False
    elif not meaningful or all(token in COMMON_PHRASE_TOKENS for token in tokens):
        return False

    requires_boundary = len(phrase) < NON_PDF_MIN_PHRASE_LEN or len(tokens) <= 2
    if requires_boundary:
        return re.search(_phrase_boundary_pattern(phrase), haystack) is not None

    return phrase in haystack


def _highlight_phrase_safe_for_output(phrase: str) -> bool:
    normalized = _normalize_text(phrase)
    if not normalized:
        return False
    tokens = _tokenize_normalized(normalized)
    if not tokens or len(normalized) == 1:
        return False
    meaningful = _meaningful_token_set(normalized)
    if len(tokens) == 1:
        return len(tokens[0]) >= 3 and tokens[0] not in COMMON_PHRASE_TOKENS
    if not meaningful or all(token in COMMON_PHRASE_TOKENS for token in tokens):
        return False
    return True


def _dedupe_keep_best(candidates: list[dict]) -> list[dict]:
    best_by_phrase: dict[str, dict] = {}
    for candidate in candidates:
        normalized = candidate.get("normalizedPhrase", "")
        if not normalized:
            continue
        current = best_by_phrase.get(normalized)
        if current is None:
            best_by_phrase[normalized] = candidate
            continue
        if float(candidate.get("score", 0.0)) > float(current.get("score", 0.0)):
            best_by_phrase[normalized] = candidate
            continue
        if float(candidate.get("score", 0.0)) == float(current.get("score", 0.0)) and int(candidate.get("priority", 0)) > int(current.get("priority", 0)):
            best_by_phrase[normalized] = candidate
    return list(best_by_phrase.values())


def _collect_non_pdf_candidates(evidence_anchor: str, evidence_context: str, evidence_text: str, snippet: str, chunk_text: str, answer: str, question: str) -> list[dict]:
    candidates: list[dict] = []

    def add_candidate(match_type: str, phrase: str, score: float, kind: str = "phrase") -> None:
        normalized = _normalize_text(phrase)
        if not normalized:
            return
        min_len = NON_PDF_MIN_PHRASE_LEN
        if len(normalized) < min_len and score < 1.0:
            return
        if _phrase_is_common(normalized) and score < 1.0:
            return
        candidates.append({
            "matchType": match_type,
            "phrase": phrase.strip(),
            "normalizedPhrase": normalized,
            "score": score,
            "kind": kind,
            "priority": SOURCE_MATCH_PRIORITY.get(match_type, 0),
        })

    def add_source(match_type: str, text: str, base_score: float) -> None:
        if not text or not text.strip():
            return
        add_candidate(match_type, text, base_score, "full")
        sentences = _split_sentences(text)
        for sentence in sentences:
            add_candidate(match_type, sentence, base_score - 0.04, "sentence")
        # Keep windows conservative and sentence-biased to avoid noisy fragments.
        for sentence in sentences:
            words = _tokenize_normalized(_normalize_text(sentence))
            for size in (14, 12, 10, 8):
                if len(words) < size:
                    continue
                for idx in range(len(words) - size + 1):
                    phrase = " ".join(words[idx:idx + size]).strip()
                    add_candidate(match_type, phrase, base_score - 0.12, "window")

    add_source("evidenceAnchor", evidence_anchor, 1.18)
    add_source("evidenceContext", evidence_context, 1.14)
    add_source("evidenceText", evidence_text, 1.05)
    add_source("snippet", snippet, 1.0)
    add_source("chunkText", chunk_text, 0.9)
    add_source("answerPhrase", answer, 0.78)
    add_source("question", question, 0.66)

    deduped = _dedupe_keep_best(candidates)
    return sorted(
        deduped,
        key=lambda candidate: (
            -float(candidate.get("score", 0.0)),
            -int(candidate.get("priority", 0)),
            -len(candidate.get("normalizedPhrase", "")),
        ),
    )


def _find_all_occurrences(haystack: str, needle: str) -> list[tuple[int, int]]:
    if not haystack or not needle:
        return []
    matches: list[tuple[int, int]] = []
    start = 0
    needle_len = len(needle)
    while start < len(haystack):
        idx = haystack.find(needle, start)
        if idx == -1:
            break
        matches.append((idx, idx + needle_len))
        start = idx + max(needle_len, 1)
    return matches


def _merge_ranges(ranges: list[tuple[int, int]]) -> list[tuple[int, int]]:
    if not ranges:
        return []
    ordered = sorted(ranges, key=lambda value: (value[0], value[1]))
    merged: list[tuple[int, int]] = [ordered[0]]
    for start, end in ordered[1:]:
        last_start, last_end = merged[-1]
        if start <= last_end:
            merged[-1] = (last_start, max(last_end, end))
        else:
            merged.append((start, end))
    return merged


def _token_overlap_score(left: str, right: str) -> float:
    left_tokens = _meaningful_token_set(left)
    right_tokens = _meaningful_token_set(right)
    if not left_tokens or not right_tokens:
        return 0.0
    overlap = len(left_tokens & right_tokens)
    return overlap / max(min(len(left_tokens), len(right_tokens)), 1)


def _block_chunk_proximity(block: dict, chunk_index: int | None, chunk_id: str | None) -> float:
    block_chunk_id = str(block.get("chunk_id") or "").strip()
    if chunk_id and block_chunk_id and chunk_id == block_chunk_id:
        return 8.0

    block_chunk_index = block.get("chunk_index")
    if block_chunk_index is None:
        block_chunk_index = block.get("blockIndex")
    if block_chunk_index is None or chunk_index is None:
        return 0.0
    try:
        distance = abs(int(block_chunk_index) - int(chunk_index))
    except (TypeError, ValueError):
        return 0.0
    if distance == 0:
        return 4.0
    if distance == 1:
        return 1.2
    if distance == 2:
        return 0.45
    return 0.0


def _chunk_match_reason(block: dict, chunk_index: int | None, chunk_id: str | None) -> str:
    block_chunk_id = str(block.get("chunk_id") or "").strip()
    if chunk_id and block_chunk_id and chunk_id == block_chunk_id:
        return "chunk_id_exact"
    block_chunk_index = block.get("chunk_index")
    if block_chunk_index is None:
        block_chunk_index = block.get("blockIndex")
    if block_chunk_index is None or chunk_index is None:
        return ""
    try:
        distance = abs(int(block_chunk_index) - int(chunk_index))
    except (TypeError, ValueError):
        return ""
    if distance == 0:
        return "chunk_index_exact"
    if distance == 1:
        return "chunk_index_distance_1"
    if distance == 2:
        return "chunk_index_distance_2"
    return ""


def _best_list_item_phrase(block: dict, evidence_text: str, snippet: str, chunk_text: str, answer: str, question: str) -> tuple[str, str, float]:
    items = [str(item).strip() for item in (block.get("items") or []) if str(item).strip()]
    if not items:
        return "", "", 0.0
    best = ("", "", -1.0)
    for item in items:
        item_norm = _normalize_text(item)
        match_type = "sentence_overlap"
        score = (
            _token_overlap_score(item, evidence_text) * 3.2
            + _token_overlap_score(item, snippet) * 2.6
            + _token_overlap_score(item, chunk_text) * 1.8
            + _token_overlap_score(item, answer) * 0.55
            + _token_overlap_score(item, question) * 0.35
        )
        for source_name, source_value, bonus in (
            ("evidenceText", evidence_text, 4.5),
            ("snippet", snippet, 3.8),
            ("chunkText", chunk_text, 2.4),
        ):
            source_norm = _normalize_text(source_value)
            if item_norm and source_norm and (
                _phrase_occurs_in_normalized_text(source_norm, item_norm)
                or _phrase_occurs_in_normalized_text(item_norm, source_norm)
            ):
                score += bonus
                match_type = source_name
                break
        if score > best[2]:
            best = (item, match_type, score)
    return best


def _best_table_phrase(block: dict, evidence_text: str, snippet: str, chunk_text: str, answer: str, question: str) -> tuple[str, str, float]:
    rows = block.get("rows") or []
    row_texts = [" | ".join(str(cell).strip() for cell in row if str(cell).strip()) for row in rows if row]
    row_texts = [row for row in row_texts if row.strip()]
    if not row_texts:
        return "", "", 0.0
    best = ("", "", -1.0)
    for row in row_texts:
        row_norm = _normalize_text(row)
        match_type = "sentence_overlap"
        score = (
            _token_overlap_score(row, evidence_text) * 3.0
            + _token_overlap_score(row, snippet) * 2.5
            + _token_overlap_score(row, chunk_text) * 1.7
            + _token_overlap_score(row, answer) * 0.9
            + _token_overlap_score(row, question) * 0.8
        )
        for source_name, source_value, bonus in (
            ("evidenceText", evidence_text, 4.2),
            ("snippet", snippet, 3.4),
            ("chunkText", chunk_text, 2.0),
        ):
            source_norm = _normalize_text(source_value)
            if row_norm and source_norm and (
                _phrase_occurs_in_normalized_text(source_norm, row_norm)
                or _phrase_occurs_in_normalized_text(row_norm, source_norm)
            ):
                score += bonus
                match_type = source_name
                break
        if score > best[2]:
            best = (row, match_type, score)
    return best


def _best_sentence_for_block(block_text: str, evidence_text: str, snippet: str, chunk_text: str, answer: str, question: str) -> tuple[str, str, float]:
    source = evidence_text or snippet or chunk_text or answer or question
    sentences = _split_sentences(block_text)
    if not sentences:
        return block_text, "block_fallback", 0.0

    best_sentence = sentences[0]
    best_score = -1.0
    for sentence in sentences:
        score = (
            _token_overlap_score(sentence, evidence_text) * 0.45
            + _token_overlap_score(sentence, snippet) * 0.32
            + _token_overlap_score(sentence, answer) * 0.16
            + _token_overlap_score(sentence, question) * 0.07
        )
        if chunk_text and _phrase_occurs_in_normalized_text(_normalize_text(chunk_text), _normalize_text(sentence)):
            score += 0.22
        if source and _normalize_text(sentence) == _normalize_text(source):
            score += 0.4
        if score > best_score:
            best_score = score
            best_sentence = sentence

    return best_sentence, "sentence_overlap" if best_score > 0 else "block_fallback", best_score


def _score_non_pdf_block(
    block: dict,
    candidates: list[dict],
    evidence_anchor: str,
    evidence_context: str,
    evidence_text: str,
    snippet: str,
    chunk_text: str,
    answer: str,
    question: str,
    chunk_index: int | None,
    chunk_id: str | None,
    fallback_index: int,
) -> dict | None:
    block_index = int(block.get("blockIndex", fallback_index) or fallback_index)
    raw_block_text = _normalize_block_text(block)
    block_text = _normalize_text(raw_block_text)
    if not block_text or block.get("type") == "heading":
        return None

    chunk_score = _block_chunk_proximity(block, chunk_index, chunk_id)
    chunk_reason = _chunk_match_reason(block, chunk_index, chunk_id)
    exact_hits: list[dict] = []

    for candidate in candidates:
        normalized_phrase = candidate.get("normalizedPhrase", "")
        if not normalized_phrase or not _phrase_occurs_in_normalized_text(block_text, normalized_phrase):
            continue
        exact_hits.append({
            "phrase": candidate.get("phrase", ""),
            "normalizedPhrase": normalized_phrase,
            "matchType": candidate.get("matchType", "snippet"),
            "score": float(candidate.get("score", 0.0)),
            "priority": int(candidate.get("priority", 0)),
            "kind": candidate.get("kind", ""),
        })

    anchor_norm = _normalize_text(evidence_anchor)
    if anchor_norm and _phrase_occurs_in_normalized_text(block_text, anchor_norm):
        exact_hits.append({
            "phrase": evidence_anchor,
            "normalizedPhrase": anchor_norm,
            "matchType": "evidenceAnchor",
            "score": 6.5,
            "priority": 10,
            "kind": "anchor",
        })

    context_norm = _normalize_text(evidence_context)
    if context_norm and _phrase_occurs_in_normalized_text(block_text, context_norm):
        exact_hits.append({
            "phrase": evidence_context,
            "normalizedPhrase": context_norm,
            "matchType": "evidenceContext",
            "score": 5.5,
            "priority": 9,
            "kind": "context",
        })

    if block.get("type") == "list":
        phrase, match_type, phrase_score = _best_list_item_phrase(block, evidence_text, snippet, chunk_text, answer, question)
        if phrase:
            exact_hits.append({
                "phrase": phrase,
                "normalizedPhrase": _normalize_text(phrase),
                "matchType": match_type,
                "score": phrase_score,
                "priority": SOURCE_MATCH_PRIORITY.get(match_type, 0),
                "kind": "list_item",
            })

    if block.get("type") == "table":
        phrase, match_type, phrase_score = _best_table_phrase(block, evidence_text, snippet, chunk_text, answer, question)
        if phrase:
            exact_hits.append({
                "phrase": phrase,
                "normalizedPhrase": _normalize_text(phrase),
                "matchType": match_type,
                "score": phrase_score,
                "priority": SOURCE_MATCH_PRIORITY.get(match_type, 0),
                "kind": "table_row",
            })

    exact_hits.sort(key=lambda hit: (
        -SOURCE_MATCH_PRIORITY.get(hit.get("matchType", ""), 0),
        -float(hit.get("score", 0.0)),
        -len(hit.get("normalizedPhrase", "")),
    ))

    phrase = ""
    match_type = ""
    phrase_score = 0.0
    if exact_hits:
        best_hit = exact_hits[0]
        phrase = best_hit.get("phrase", "")
        match_type = best_hit.get("matchType", "snippet")
        phrase_score = float(best_hit.get("score", 0.0))
    else:
        phrase, match_type, phrase_score = _best_sentence_for_block(raw_block_text, evidence_text, snippet, chunk_text, answer, question)

    evidence_overlap = _token_overlap_score(raw_block_text, evidence_text)
    snippet_overlap = _token_overlap_score(raw_block_text, snippet)
    chunk_overlap = _token_overlap_score(raw_block_text, chunk_text)
    answer_overlap = _token_overlap_score(raw_block_text, answer)
    question_overlap = _token_overlap_score(raw_block_text, question)

    exact_bonus = 0.0
    if match_type == "evidenceText":
        exact_bonus = 5.0
    elif match_type == "snippet":
        exact_bonus = 4.0
    elif match_type == "chunkText":
        exact_bonus = 1.8

    fallback_score = 0.0
    if not exact_hits:
        fallback_score = (
            phrase_score * 1.4
            + chunk_overlap * 1.2
            + evidence_overlap * 1.0
            + snippet_overlap * 0.9
            + answer_overlap * 0.3
            + question_overlap * 0.2
        )

    type_bonus = 0.0
    question_norm = _normalize_text(question)
    if block.get("type") == "table" and re.search(r"\b(table|matrix|status|approver|owner|value)\b", question_norm):
        type_bonus = 1.0
    elif block.get("type") == "list" and re.search(r"\b(list|requirement|must|should)\b", question_norm):
        type_bonus = 0.45

    score = (
        chunk_score
        + exact_bonus
        + phrase_score
        + evidence_overlap * 2.2
        + snippet_overlap * 1.8
        + fallback_score
        + type_bonus
    )

    if score < 0.35:
        return None

    reason_parts = [part for part in [
        chunk_reason,
        f"{match_type}_match" if match_type else "",
        "type_hint" if type_bonus else "",
        "fallback_overlap" if fallback_score else "",
    ] if part]

    return {
        "blockIndex": block_index,
        "blockChunkIndex": block.get("chunk_index") if block.get("chunk_index") is not None else block.get("blockIndex"),
        "blockChunkId": block.get("chunk_id"),
        "score": score,
        "hits": exact_hits,
        "phrase": phrase,
        "matchType": match_type or ("chunk_proximity" if chunk_score else "sentence_overlap"),
        "reason": ", ".join(reason_parts) or "token_overlap",
        "preview": raw_block_text[:220],
        "chunkScore": chunk_score,
    }


def _select_non_pdf_highlights(
    preview_blocks: list[dict],
    candidates: list[dict],
    evidence_anchor: str = "",
    evidence_context: str = "",
    evidence_text: str = "",
    snippet: str = "",
    chunk_text: str = "",
    answer: str = "",
    question: str = "",
    chunk_index: int | None = None,
    chunk_id: str | None = None,
) -> tuple[list[dict], list[int], list[int], dict]:
    block_scores: list[dict] = []

    for idx, block in enumerate(preview_blocks):
        scored = _score_non_pdf_block(
            block,
            candidates,
            evidence_anchor,
            evidence_context,
            evidence_text,
            snippet,
            chunk_text,
            answer,
            question,
            chunk_index,
            chunk_id,
            idx,
        )
        if scored:
            block_scores.append(scored)

    ranked_blocks = sorted(block_scores, key=lambda item: (-float(item["score"]), item["blockIndex"]))
    selected_blocks = ranked_blocks[:NON_PDF_BLOCK_LIMIT]
    debug = {
        "receivedChunkIndex": chunk_index,
        "receivedChunkId": chunk_id,
        "topScoredBlocks": [
            {
                "blockIndex": item["blockIndex"],
                "score": round(float(item["score"]), 4),
                "matchType": item.get("matchType"),
                "phrase": item.get("phrase"),
                "reason": item.get("reason"),
                "blockChunkIndex": item.get("blockChunkIndex"),
                "blockChunkId": item.get("blockChunkId"),
                "preview": item.get("preview", ""),
            }
            for item in ranked_blocks[:5]
        ],
    }
    if not selected_blocks:
        return [], [], [], debug

    primary = selected_blocks[0]
    primary_blocks = [primary["blockIndex"]]
    secondary_blocks = [
        item["blockIndex"]
        for item in selected_blocks[1:NON_PDF_BLOCK_LIMIT]
        if abs(int(item["blockIndex"]) - int(primary["blockIndex"])) <= 2 or item.get("chunkScore", 0) > 0
    ][:2]

    matches: list[dict] = []
    if primary.get("phrase") and _highlight_phrase_safe_for_output(str(primary.get("phrase") or "")):
        matches.append({
            "pageNumber": None,
            "blockIndex": primary["blockIndex"],
            "phrase": primary.get("phrase", ""),
            "matchType": primary.get("matchType", "sentence_overlap"),
            "score": round(float(primary.get("score", 0.0)), 4),
            "isPrimary": True,
            "reason": primary.get("reason", ""),
            "blockChunkIndex": primary.get("blockChunkIndex"),
            "blockChunkId": primary.get("blockChunkId"),
        })

    debug.update({
        "selectedPrimaryBlockIndex": primary_blocks[0],
        "selectedSecondaryBlockIndexes": secondary_blocks,
        "matchType": matches[0].get("matchType") if matches else "",
        "phrase": matches[0].get("phrase") if matches else "",
        "score": matches[0].get("score") if matches else 0,
        "reason": matches[0].get("reason") if matches else "",
        "matchedBlockChunkIndex": primary.get("blockChunkIndex"),
        "matchedBlockChunkId": primary.get("blockChunkId"),
    })

    return matches, primary_blocks, secondary_blocks, debug


def _normalize_block_text(block: dict) -> str:
    block_type = block.get("type")
    if block_type == "table":
        rows = block.get("rows") or []
        return "\n".join(" | ".join(str(cell) for cell in row if str(cell).strip()) for row in rows if row)
    if block_type == "list":
        items = block.get("items") or []
        return "\n".join(str(item) for item in items if str(item).strip())
    return str(block.get("text") or "")


def _iter_docx_body_items(doc: DocxDocument):
    for child in doc.element.body.iterchildren():
        if isinstance(child, CT_P):
            yield Paragraph(child, doc)
        elif isinstance(child, CT_Tbl):
            yield Table(child, doc)


def _docx_block_from_paragraph(paragraph: Paragraph) -> dict | None:
    text = _clean_preview_text(paragraph.text or "")
    if not text:
        return None

    style_name = ""
    try:
        style_name = (paragraph.style.name or "").strip()
    except Exception:
        style_name = ""

    style_lower = style_name.lower()
    heading = re.match(r"heading\s+(\d+)", style_lower)
    if heading:
        return {"type": "heading", "level": min(max(int(heading.group(1)), 1), 3), "text": text}

    if style_lower in {"title", "subtitle"}:
        return {"type": "heading", "level": 1 if style_lower == "title" else 2, "text": text}

    if any(token in style_lower for token in ("list", "bullet", "number")):
        return {"type": "list", "level": None, "items": [text], "text": text}

    if _looks_like_heading(text):
        return {"type": "heading", "level": 2, "text": text}

    if _is_structured_line(text):
        return {"type": "list", "level": None, "items": [text], "text": text}

    return {"type": "paragraph", "level": None, "text": text}


def _docx_block_from_table(table: Table) -> dict | None:
    rows: list[list[str]] = []
    for row in table.rows:
        cells = [_clean_preview_text(cell.text or "") for cell in row.cells]
        cells = [cell for cell in cells if cell]
        if cells:
            rows.append(cells)

    if not rows:
        return None

    flat_text = "\n".join(" | ".join(row) for row in rows)
    return {"type": "table", "level": None, "text": flat_text, "rows": rows}


def _apply_section_context(blocks: list[dict]) -> list[dict]:
    heading_stack: list[tuple[int, str]] = []

    for idx, block in enumerate(blocks):
        block["blockIndex"] = idx
        block.setdefault("page_number", None)
        block.setdefault("chunk_id", None)
        block.setdefault("chunk_index", None)

        block_type = block.get("type") or "paragraph"
        block["type"] = block_type if block_type in {"heading", "paragraph", "list", "table", "pre", "fallback"} else "paragraph"
        current_section = heading_stack[-1][1] if heading_stack else None
        parent_section = heading_stack[-2][1] if len(heading_stack) > 1 else None
        block.setdefault("section_title", current_section)
        block.setdefault("parent_section_title", parent_section)

        if block["type"] == "heading":
            level = int(block.get("level") or 2)
            title = _clean_preview_text(str(block.get("text") or ""))
            while heading_stack and heading_stack[-1][0] >= level:
                heading_stack.pop()
            parent_section = heading_stack[-1][1] if heading_stack else None
            if title:
                heading_stack.append((level, title))
            block["section_title"] = title or current_section
            block["parent_section_title"] = parent_section

    return blocks


def _merge_adjacent_list_blocks(blocks: list[dict]) -> list[dict]:
    merged: list[dict] = []
    for block in blocks:
        if block.get("type") == "list" and merged and merged[-1].get("type") == "list":
            previous = merged[-1]
            previous_items = previous.setdefault("items", [])
            previous_items.extend(block.get("items") or [block.get("text", "")])
            previous["text"] = "\n".join(str(item) for item in previous_items if str(item).strip())
            continue
        merged.append(block)
    return merged


@lru_cache(maxsize=64)
def _extract_pdf_text_layer_cached(path_str: str, mtime_ns: int, file_size: int) -> tuple[dict, ...]:
    from pypdf import PdfReader

    reader = PdfReader(path_str)
    pages: list[dict] = []

    for page_number, page in enumerate(reader.pages, start=1):
        raw_text = page.extract_text() or ""
        lines = [line.rstrip() for line in raw_text.splitlines()]
        rebuilt: list[str] = []
        for line in lines:
            stripped = line.strip()
            if not stripped:
                if rebuilt and rebuilt[-1] != "":
                    rebuilt.append("")
                continue
            if rebuilt and rebuilt[-1] and not re.search(r"[.!?:]\s*$", rebuilt[-1]):
                rebuilt[-1] = f"{rebuilt[-1]} {stripped}"
            else:
                rebuilt.append(stripped)

        page_text = "\n".join(rebuilt).strip()
        pages.append({
            "pageNumber": page_number,
            "text": page_text,
        })

    return tuple(pages)


@lru_cache(maxsize=64)
def _extract_docx_preview_cached(path_str: str, mtime_ns: int, file_size: int) -> dict:
    doc = DocxDocument(path_str)
    blocks: list[dict] = []
    plain_text_parts: list[str] = []

    for item in _iter_docx_body_items(doc):
        block: dict | None
        if isinstance(item, Paragraph):
            block = _docx_block_from_paragraph(item)
        else:
            block = _docx_block_from_table(item)
        if not block:
            continue
        blocks.append(block)

    blocks = _apply_section_context(_merge_adjacent_list_blocks(blocks))
    for block in blocks:
        block_text = _normalize_block_text(block)
        if block_text:
            plain_text_parts.append(block_text)

    return {
        "documentTextLayer": blocks,
        "plainText": _clean_preview_text("\n\n".join(plain_text_parts)),
        "extraction_status": "ok" if blocks else "empty",
    }


@lru_cache(maxsize=64)
def _extract_text_preview_cached(path_str: str, mtime_ns: int, file_size: int) -> dict:
    file_path = Path(path_str)
    ext = file_path.suffix.lower()
    raw = file_path.read_text(encoding="utf-8", errors="ignore")
    raw = raw.replace("\r\n", "\n").replace("\r", "\n")

    if ext == ".rtf":
        raw = re.sub(r"\\par[d]?", "\n", raw)
        raw = re.sub(r"\\tab", "\t", raw)
        raw = re.sub(r"\\'[0-9a-fA-F]{2}", "", raw)
        raw = re.sub(r"\\[a-zA-Z]+-?\d*\s?", "", raw)
        raw = raw.replace("{", "").replace("}", "")

    raw = _clean_preview_text(raw)
    blocks: list[dict] = []

    if not raw:
        return {"documentTextLayer": [], "plainText": "", "extraction_status": "empty"}

    groups = _split_paragraph_groups(raw)
    for group in groups:
        joined = _clean_preview_text("\n".join(group))
        if not joined:
            continue

        block_type = "paragraph"
        level = None
        text_value = joined
        items = None
        rows = None

        if ext == ".md":
            first = group[0].strip()
            is_heading, heading_level = _is_markdown_heading(first)
            if is_heading:
                block_type = "heading"
                level = heading_level
                text_value = re.sub(r"^#{1,6}\s+", "", first).strip()
            elif _is_markdown_table_group(group):
                rows = []
                for row_line in group:
                    stripped = row_line.strip()
                    if not stripped or re.match(r"^\s*\|?\s*[:\-\| ]+\|?\s*$", stripped):
                        continue
                    cells = _split_markdown_table_row(stripped)
                    if cells:
                        rows.append(cells)
                if rows:
                    block_type = "table"
                    text_value = "\n".join(" | ".join(row) for row in rows)
            elif any(_is_markdown_list_item(line) for line in group):
                block_type = "list"
                items = [line.strip().lstrip("-*+").strip() for line in group if line.strip()]
            elif first.startswith("```") and joined.endswith("```"):
                block_type = "pre"
                code_lines = list(group)
                if code_lines and code_lines[0].strip().startswith("```"):
                    code_lines = code_lines[1:]
                if code_lines and code_lines[-1].strip() == "```":
                    code_lines = code_lines[:-1]
                text_value = "\n".join(code_lines).strip("\n")
            elif _looks_like_heading(first):
                block_type = "heading"
                level = 2
                text_value = first
        else:
            if len(group) == 1:
                line = group[0].strip()
                if _looks_like_heading(line):
                    block_type = "heading"
                    level = 2
                elif _is_structured_line(line):
                    block_type = "list"
                    items = [line]
                text_value = line
            else:
                if any(_is_structured_line(line) for line in group):
                    block_type = "list"
                    items = [line.strip() for line in group if line.strip()]
                    text_value = "\n".join(items)

        block: dict = {"type": block_type, "level": level, "text": text_value}
        if items is not None:
            block["items"] = items
        if rows is not None:
            block["rows"] = rows
        blocks.append(block)

    blocks = _apply_section_context(blocks)
    plain_text = "\n\n".join(_normalize_block_text(block) for block in blocks if _normalize_block_text(block))
    return {"documentTextLayer": blocks, "plainText": _clean_preview_text(plain_text), "extraction_status": "ok" if blocks else "empty"}


def _extract_preview_payload(file_path: Path, stat: os.stat_result, *, include_pdf_text_layer: bool = True) -> dict:
    """
    Build the preview payload for a document.

    For PDFs, `include_pdf_text_layer=False` skips the expensive full-page
    text extraction and just returns an empty text layer. This is used by
    /document/meta which only needs page count + file metadata — the full
    text layer is extracted later (and cached) by /document/highlight when
    actual highlighting is requested.
    """
    ext = file_path.suffix.lower().lstrip(".")
    if ext == "pdf":
        if include_pdf_text_layer:
            return {
                "pdfTextLayer": list(_extract_pdf_text_layer_cached(str(file_path), stat.st_mtime_ns, stat.st_size)),
                "documentTextLayer": [],
                "plainText": "",
                "extraction_status": "ok",
            }
        # Fast path: just page count, no text extraction.
        return {
            "pdfTextLayer": [],
            "pdfPageCount": _pdf_page_count_fast(str(file_path), stat.st_mtime_ns, stat.st_size),
            "documentTextLayer": [],
            "plainText": "",
            "extraction_status": "ok",
        }
    if ext == "docx":
        return _extract_docx_preview_cached(str(file_path), stat.st_mtime_ns, stat.st_size)
    if ext in {"xlsx", "xlsm"}:
        return _xlsx_sheet_payload(file_path)
    if ext == "csv":
        return _csv_sheet_payload(file_path)
    if ext in {"pptx", "ppt"}:
        return {"documentTextLayer": [], "plainText": "", "extraction_status": "unsupported:ppt_disabled"}
    if ext == "xls":
        return {"documentTextLayer": [], "plainText": "", "sheets": [], "extraction_status": "unsupported:xls"}
    if ext == "ods":
        return {"documentTextLayer": [], "plainText": "", "sheets": [], "extraction_status": "unsupported:ods"}
    if ext in {"txt", "md", "rtf"}:
        return _extract_text_preview_cached(str(file_path), stat.st_mtime_ns, stat.st_size)
    return {"documentTextLayer": [], "plainText": "", "extraction_status": "unsupported"}


def build_document_metadata(path: str, *, include_text_layer: bool = False) -> dict:
    """
    Build metadata payload for a document.

    By default (`include_text_layer=False`) this uses a fast path for PDFs
    that only reads page count without extracting any text. This makes
    /document/meta respond in ~10ms on first hit instead of ~1s for large
    PDFs. When highlighting is needed, build_document_highlights() calls
    this with include_text_layer=True (or extracts the text layer directly).
    """
    perf_start = time.perf_counter()
    file_path = resolve_document_path(path)
    stat = file_path.stat()
    ext = file_path.suffix.lower().lstrip(".")
    content_type = _normalize_content_type(ext, mimetypes.guess_type(file_path.name)[0] or "application/octet-stream")
    metadata = {
        "file_name": file_path.name,
        "source_path": str(file_path.relative_to(INTAKE_FOLDER)).replace("\\", "/"),
        "extension": ext,
        "file_size": stat.st_size,
        "last_modified": int(stat.st_mtime),
        "last_modified_iso": datetime.fromtimestamp(stat.st_mtime).isoformat(),
        "content_type": content_type,
        "page_count": None,
        "viewerType": "unsupported",
        "canPreview": False,
        "pdfTextLayer": [],
        "documentTextLayer": [],
        "spreadsheetSheets": [],
        "targetSheet": None,
        "plainText": "",
        "extraction_status": "pending",
    }

    preview = _extract_preview_payload(file_path, stat, include_pdf_text_layer=include_text_layer)

    if ext == "pdf":
        metadata["content_type"] = "application/pdf"
        # Fast path populates pdfPageCount; slow path uses len(pdfTextLayer).
        if include_text_layer:
            metadata["page_count"] = len(preview["pdfTextLayer"])
            metadata["pdfTextLayer"] = preview["pdfTextLayer"]
        else:
            metadata["page_count"] = int(preview.get("pdfPageCount", 0) or 0)
            # pdfTextLayer stays [] — frontend doesn't need it for metadata,
            # PDF.js inside the iframe reads the PDF directly for rendering.
        metadata["viewerType"] = "pdf"
        metadata["canPreview"] = True
        metadata["extraction_status"] = preview.get("extraction_status", "ok")
        _perf_log("build_document_metadata", perf_start, ext=ext, path=file_path.name, include_text_layer=include_text_layer)
        return metadata

    if ext == "docx":
        metadata["viewerType"] = "docx"
        metadata["canPreview"] = True
        metadata["documentTextLayer"] = preview.get("documentTextLayer", [])
        metadata["plainText"] = preview.get("plainText", "")
        metadata["extraction_status"] = preview.get("extraction_status", "ok" if metadata["plainText"] else "empty")
        _perf_log("build_document_metadata", perf_start, ext=ext, path=file_path.name)
        return metadata

    if ext in {"xlsx", "xlsm", "csv"}:
        metadata["viewerType"] = "xlsx"
        metadata["canPreview"] = True
        metadata["spreadsheetSheets"] = preview.get("sheets", [])
        metadata["targetSheet"] = metadata["spreadsheetSheets"][0] if metadata["spreadsheetSheets"] else None
        metadata["extraction_status"] = preview.get("extraction_status", "ok" if metadata["spreadsheetSheets"] else "empty")
        _perf_log("build_document_metadata", perf_start, ext=ext, path=file_path.name)
        return metadata

    if ext in {"pptx", "ppt"}:
        metadata["viewerType"] = "unsupported"
        metadata["canPreview"] = False
        metadata["extraction_status"] = "unsupported:ppt_disabled"
        return metadata

    if ext in {"txt", "md", "rtf"}:
        metadata["viewerType"] = "text"
        metadata["canPreview"] = True
        metadata["documentTextLayer"] = preview.get("documentTextLayer", [])
        metadata["plainText"] = preview.get("plainText", "")
        metadata["extraction_status"] = preview.get("extraction_status", "ok" if metadata["plainText"] else "empty")
        _perf_log("build_document_metadata", perf_start, ext=ext, path=file_path.name)
        return metadata

    metadata["extraction_status"] = preview.get("extraction_status", "unsupported")
    return metadata


def build_view_source_payload(path: str, snippet: str = "") -> dict:
    file_path = resolve_document_path(path)
    metadata = build_document_metadata(path)
    viewer_type = metadata.get("viewerType", "unsupported")

    payload = {
        "filename": file_path.name,
        "snippet": snippet,
        "path": str(file_path),
        "type": viewer_type,
        "meta": metadata,
    }

    if viewer_type == "pdf":
        payload["pages"] = metadata.get("pdfTextLayer", [])
        payload["text"] = "\n\n".join(page.get("text", "") for page in metadata.get("pdfTextLayer", []))
    elif viewer_type == "xlsx":
        payload["workbook"] = file_path.name
        payload["sheets"] = metadata.get("spreadsheetSheets", [])
        payload["targetSheet"] = metadata.get("targetSheet")
        payload["text"] = "\n\n".join(
            "\n".join(str(cell.get("value", "")) for cell in row.get("cells", []))
            for sheet in metadata.get("spreadsheetSheets", [])
            for row in sheet.get("preview_rows", [])
        )
    else:
        payload["blocks"] = metadata.get("documentTextLayer", [])
        payload["text"] = metadata.get("plainText", "")

    return payload


def serve_document_file(path: str) -> FileResponse:
    file_path = resolve_document_path(path)
    content_type = mimetypes.guess_type(file_path.name)[0] or "application/octet-stream"
    if file_path.suffix.lower() == ".pdf":
        content_type = "application/pdf"
    return FileResponse(path=str(file_path), media_type=content_type, filename=file_path.name)


def serve_document_inline(path: str) -> FileResponse:
    file_path = resolve_document_path(path)

    if file_path.suffix.lower() != ".pdf":
        raise HTTPException(status_code=415, detail="Inline preview only supports PDFs")

    safe_name = file_path.name.replace('"', "")
    encoded_name = quote(safe_name)

    return FileResponse(
        path=str(file_path),
        media_type="application/pdf",
        filename=safe_name,
        headers={
            "Content-Disposition": f"inline; filename=\"{safe_name}\"; filename*=UTF-8''{encoded_name}",
            "X-Content-Type-Options": "nosniff",
        },
    )


def build_document_highlights(
    path: str,
    evidence_text: str = "",
    snippet: str = "",
    chunk_text: str = "",
    answer: str = "",
    question: str = "",
    evidence_anchor: str = "",
    evidence_context: str = "",
    chunk_index: int = 0,
    chunk_id: str = "",
    page_number: int | None = None,
    extraction_method: str | None = None,
    has_text_layer: bool | None = None,
    ocr_confidence: float | None = None,
) -> dict:
    perf_start = time.perf_counter()
    file_path = resolve_document_path(path)
    # include_text_layer=True: highlighting needs the extracted PDF/doc text.
    # The first call to /document/meta was cheap (fast page-count path);
    # this is the only place that pays the full-extraction cost, and it's
    # cached so subsequent highlight requests on the same file are fast.
    metadata = build_document_metadata(path, include_text_layer=True)
    viewer_type = metadata.get("viewerType")
    is_pdf = viewer_type == "pdf"
    preview_blocks = metadata.get("pdfTextLayer", []) if is_pdf else metadata.get("documentTextLayer", [])
    is_xlsx = viewer_type == "xlsx"
    xlsx_sheets = metadata.get("spreadsheetSheets", []) if is_xlsx else []
    resolved_extraction_method = (extraction_method or "").strip().lower() or None
    if resolved_extraction_method not in {"text_layer", "ocr", "ocr_failed", "vision_fallback"}:
        resolved_extraction_method = None
    resolved_has_text_layer = has_text_layer
    if resolved_has_text_layer is None:
        if resolved_extraction_method == "text_layer":
            resolved_has_text_layer = True
        elif resolved_extraction_method in {"ocr", "ocr_failed"}:
            resolved_has_text_layer = False
    resolved_ocr_confidence = ocr_confidence
    requested_path = str(path or "")
    requested_chunk_id = str(chunk_id or "").strip()
    requested_chunk_index = int(chunk_index or 0) if chunk_index is not None else 0
    requested_page_number = int(page_number) if page_number is not None else None

    _viewer_debug(
        "highlight request",
        path=requested_path,
        viewer_type=viewer_type,
        chunk_id=requested_chunk_id,
        chunk_index=requested_chunk_index,
        page_number=requested_page_number,
    )

    def _response_payload(
        *,
        matches: list[dict] | None = None,
        normalized_length: int = 0,
        phrase_preview: str = "",
        candidate_count: int = 0,
        primary_block_indexes: list[int] | None = None,
        secondary_block_indexes: list[int] | None = None,
        debug: dict | None = None,
        warning: str | None = None,
        forced_found: bool | None = None,
        forced_match_type: str | None = None,
        forced_page_number: int | None = None,
    ) -> dict:
        all_matches = list(matches or [])
        primary = all_matches[0] if all_matches else {}
        found = bool(primary.get("phrase")) if forced_found is None else bool(forced_found)
        resolved_page = forced_page_number
        if resolved_page is None:
            resolved_page = primary.get("pageNumber")
        if resolved_page is None:
            resolved_page = requested_page_number

        match_type = forced_match_type or str(primary.get("matchType") or ("no_match" if not found else ""))
        payload = {
            "found": found,
            "path": requested_path,
            "pageNumber": resolved_page,
            "chunkId": requested_chunk_id,
            "chunkIndex": requested_chunk_index,
            "matchType": match_type,
            "matchedText": str(primary.get("phrase") or ""),
            "primaryMatchText": str(primary.get("phrase") or ""),
            "evidenceAnchor": evidence_anchor or "",
            "evidenceContext": evidence_context or "",
            "evidenceText": evidence_text or "",
            "snippet": snippet or "",
            "highlightTerms": [m.get("phrase") for m in all_matches if m.get("phrase")][:8],
            "candidates": [m.get("phrase") for m in all_matches if m.get("phrase")][:8],
            "confidence": float(primary.get("score", 0.0) or 0.0),
            "score": float(primary.get("score", 0.0) or 0.0),
            "matches": all_matches,
            "primaryBlockIndexes": list(primary_block_indexes or []),
            "secondaryBlockIndexes": list(secondary_block_indexes or []),
            "normalizedLength": int(normalized_length or 0),
            "phrasePreview": phrase_preview or "",
            "candidateCount": int(candidate_count or 0),
        }
        if warning:
            payload["warning"] = warning
        if debug:
            payload["debug"] = debug

        _viewer_debug(
            "highlight response",
            found=payload["found"],
            match_type=payload["matchType"],
            page_number=payload["pageNumber"],
            matched_text_len=len(payload["matchedText"]),
            candidate_count=payload["candidateCount"],
            warning=payload.get("warning"),
        )
        _perf_log(
            "build_document_highlights",
            perf_start,
            path=file_path.name,
            viewer=viewer_type,
            found=payload["found"],
            match_type=payload["matchType"],
        )
        return payload

    def _build_pdf_fallback_payload() -> dict:
        fallback_text = _clean_preview_text(evidence_anchor or evidence_context or evidence_text or snippet or chunk_text or answer or question)
        if not fallback_text:
            return _response_payload(
                matches=[],
                normalized_length=0,
                phrase_preview="",
                candidate_count=0,
                warning="No highlightable text found in the PDF text layer.",
                forced_found=False,
                forced_match_type="no_match",
            )
        fallback_page = int(page_number or 1)
        return _response_payload(
            matches=[],
            normalized_length=len(_normalize_text(fallback_text)),
            phrase_preview=fallback_text,
            candidate_count=1,
            warning="Relevant page found, but exact PDF text highlight was not available.",
            forced_found=False,
            forced_match_type="no_match",
            forced_page_number=fallback_page,
        )

    if is_xlsx:
        query = _clean_preview_text(evidence_text or snippet or chunk_text or answer or question)
        question_norm = _normalize_text(question)
        matches: list[dict] = []
        best_sheet: dict | None = None
        best_row: dict | None = None
        best_score = -1.0

        for sheet in xlsx_sheets:
            sheet_name = str(sheet.get("sheet_name") or "")
            used_range = str(sheet.get("used_range") or "")
            headers = [str(value) for value in sheet.get("headers", []) if str(value).strip()]
            sheet_norm = _normalize_text(sheet_name)
            score = 0.0
            match_type = "chunk_proximity"
            if sheet_name and sheet_norm in question_norm:
                score += 2.0
                match_type = "exact_range"
            if used_range and used_range.lower() in question_norm:
                score += 2.5
                match_type = "exact_range"
            if headers and any(header.lower() in question_norm for header in headers):
                score += 1.2
                if match_type == "chunk_proximity":
                    match_type = "header_value_match"
            for row in sheet.get("preview_rows", []):
                row_text = " ".join(str(cell.get("value", "")) for cell in row.get("cells", []))
                row_norm = _normalize_text(row_text)
                row_score = score
                if query and query.lower() in row_norm:
                    row_score += 2.0
                    match_type = "exact_range"
                if any(token and token in row_norm for token in _normalize_text(question).split() if len(token) >= 3):
                    row_score += 0.4
                if row_score > best_score:
                    best_score = row_score
                    best_sheet = sheet
                    best_row = row

        if best_sheet:
            row_number = best_row.get("row_number") if best_row else None
            used_range = str(best_sheet.get("used_range") or "")
            if row_number:
                matches.append({
                    "sheetName": best_sheet.get("sheet_name"),
                    "rangeRef": _xlsx_range_ref(row_number, row_number, 1, max(len(best_sheet.get("headers", [])), 1)),
                    "rowNumber": row_number,
                    "phrase": query or best_sheet.get("sheet_name") or "",
                    "matchType": "row_overlap" if match_type != "exact_range" else "exact_range",
                    "score": round(max(best_score, 0.1), 4),
                })
            elif used_range:
                matches.append({
                    "sheetName": best_sheet.get("sheet_name"),
                    "rangeRef": used_range,
                    "phrase": query or best_sheet.get("sheet_name") or "",
                    "matchType": match_type,
                    "score": round(max(best_score, 0.1), 4),
                })

        return _response_payload(
            matches=matches,
            normalized_length=len(query),
            phrase_preview=query or (matches[0]["phrase"] if matches else ""),
            candidate_count=len(xlsx_sheets),
            debug={
                "sheet_name": best_sheet.get("sheet_name") if best_sheet else None,
                "range_ref": best_sheet.get("used_range") if best_sheet else None,
                "reason_selected": matches[0]["matchType"] if matches else "chunk_proximity",
            },
        )

    if not preview_blocks:
        if is_pdf:
            return _build_pdf_fallback_payload()
        return _response_payload(matches=[], normalized_length=0, phrase_preview="", candidate_count=0)

    if not is_pdf:
        non_pdf_candidates = _collect_non_pdf_candidates(evidence_anchor, evidence_context, evidence_text, snippet, chunk_text, answer, question)
        if not non_pdf_candidates:
            return {
                "matches": [],
                "primaryBlockIndexes": [],
                "secondaryBlockIndexes": [],
                "normalizedLength": 0,
                "phrasePreview": "",
                "candidateCount": 0,
            }
        matches, primary_blocks, secondary_blocks, debug = _select_non_pdf_highlights(
            preview_blocks,
            non_pdf_candidates,
            evidence_anchor=evidence_anchor,
            evidence_context=evidence_context,
            evidence_text=evidence_text,
            snippet=snippet,
            chunk_text=chunk_text,
            answer=answer,
            question=question,
            chunk_index=chunk_index,
            chunk_id=chunk_id,
        )
        normalized_lengths = [len(candidate.get("normalizedPhrase", "")) for candidate in non_pdf_candidates]
        return _response_payload(
            matches=matches,
            primary_block_indexes=primary_blocks,
            secondary_block_indexes=secondary_blocks,
            normalized_length=max(normalized_lengths) if normalized_lengths else 0,
            phrase_preview=matches[0]["phrase"] if matches else (non_pdf_candidates[0]["phrase"] if non_pdf_candidates else ""),
            candidate_count=len(non_pdf_candidates),
            debug=debug,
        )

    target_page = int(page_number) if page_number is not None else None
    pdf_queries = _collect_pdf_queries(evidence_anchor, evidence_context, evidence_text, snippet, chunk_text, answer, question)
    if not pdf_queries:
        return _response_payload(
            matches=[],
            normalized_length=0,
            phrase_preview="",
            candidate_count=0,
            warning="No highlightable text found for this source.",
            forced_found=False,
            forced_match_type="no_match",
            forced_page_number=target_page,
        )

    pdf_matches = _select_pdf_highlights(preview_blocks, pdf_queries, target_page)
    normalized_lengths = [len(query.get("normalizedPhrase", "")) for query in pdf_queries]
    if pdf_matches:
        return _response_payload(
            matches=pdf_matches,
            normalized_length=max(normalized_lengths) if normalized_lengths else 0,
            phrase_preview=pdf_matches[0]["phrase"],
            candidate_count=len(pdf_queries),
        )

    # Last-resort exact path for unusual PDFs where block scoring finds no usable overlap.
    if is_pdf:
        return _build_pdf_fallback_payload()

    candidates = _collect_candidates(evidence_anchor, evidence_context, evidence_text, snippet, chunk_text, answer, question)
    if not candidates:
        return _response_payload(
            matches=[],
            normalized_length=0,
            phrase_preview="",
            candidate_count=len(pdf_queries),
            warning="No highlightable text found for this source.",
            forced_found=False,
            forced_match_type="no_match",
            forced_page_number=target_page,
        )

    ordered_candidates = sorted(
        candidates,
        key=lambda candidate: (-float(candidate.get("score", 0.0)), -len(_normalize_text(candidate.get("phrase", ""))))
    )

    matches: list[dict] = []
    seen: set[tuple[int | None, str, str]] = set()
    fallback_normalized_lengths = [len(_normalize_text(candidate["phrase"])) for candidate in ordered_candidates if candidate.get("phrase")]

    for candidate in ordered_candidates:
        phrase = candidate.get("phrase", "")
        normalized_phrase = _normalize_text(phrase)
        if not normalized_phrase:
            continue

        best_match: dict | None = None
        best_score = -1.0

        for idx, block in enumerate(preview_blocks):
            block_text = _normalize_text(_normalize_block_text(block))
            if not _phrase_occurs_in_normalized_text(block_text, normalized_phrase):
                continue

            score = float(candidate.get("score", 0.0))
            if is_pdf:
                current_page_number = int(block.get("pageNumber", 0) or 0)
                if target_page is not None:
                    proximity = max(0.0, 1.0 - abs(current_page_number - target_page) / max(len(preview_blocks), 1))
                    score += proximity * 0.08
                match = {
                    "pageNumber": current_page_number,
                    "phrase": phrase,
                    "matchType": candidate.get("matchType", "snippet"),
                    "score": round(score, 4),
                    "extractionMethod": resolved_extraction_method or "text_layer",
                    "hasTextLayer": True if resolved_has_text_layer is None else bool(resolved_has_text_layer),
                    "ocrConfidence": resolved_ocr_confidence,
                }
                if score > best_score:
                    best_score = score
                    best_match = match
            else:
                block_index = int(block.get("blockIndex", idx) or idx)
                match = {
                    "pageNumber": None,
                    "blockIndex": block_index,
                    "phrase": phrase,
                    "matchType": candidate.get("matchType", "snippet"),
                    "score": round(score, 4),
                    "extractionMethod": resolved_extraction_method or "unknown",
                    "hasTextLayer": resolved_has_text_layer,
                    "ocrConfidence": resolved_ocr_confidence,
                }
                if score > best_score:
                    best_score = score
                    best_match = match

        if not best_match:
            continue

        key = (
            best_match.get("pageNumber") if is_pdf else best_match.get("blockIndex"),
            normalized_phrase,
            best_match["matchType"],
        )
        if key in seen:
            continue
        seen.add(key)
        matches.append(best_match)

        if len(matches) >= 8:
            break

    return _response_payload(
        matches=matches,
        normalized_length=max(fallback_normalized_lengths or normalized_lengths) if (fallback_normalized_lengths or normalized_lengths) else 0,
        phrase_preview=matches[0]["phrase"] if matches else (ordered_candidates[0]["phrase"] if ordered_candidates else ""),
        candidate_count=len(pdf_queries) + len(ordered_candidates),
        warning=(
            "Relevant page found, but exact PDF text highlight was not available."
            if is_pdf and not matches and target_page is not None
            else None
        ),
        forced_found=False if (is_pdf and not matches and target_page is not None) else None,
        forced_match_type="no_match" if (is_pdf and not matches and target_page is not None) else None,
        forced_page_number=target_page if (is_pdf and target_page is not None) else None,
    )